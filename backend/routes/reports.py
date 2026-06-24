"""
routes/reports.py
Reportistica e export rendiconti (CSV/Excel/PDF).
"""
import csv
import io
from datetime import datetime

from bson import ObjectId
from fastapi import APIRouter, HTTPException, Request
from fastapi.responses import StreamingResponse

from core.auth import get_current_user
from core.db import db
from core.roles import user_has_any_role

router = APIRouter()


@router.get("/reports/rimborsi-annuali")
async def get_report_rimborsi_annuali(request: Request, anno: int):
    user = await get_current_user(request)

    if not user_has_any_role(user, ["admin", "cassiere", "superadmin", "superuser"]):
        raise HTTPException(status_code=403, detail="Permessi insufficienti")

    query = {"data": {"$regex": f"^{anno}"}}
    if not user_has_any_role(user, ["superadmin", "superuser"]):
        query["sede_id"] = user.get("sede_id")

    pipeline = [
        {"$match": query},
        {"$group": {
            "_id": "$user_id",
            "totale_rimborsi": {"$sum": 1},
            "totale_importo": {"$sum": "$importo_totale"},
            "totale_km": {"$sum": "$km_totali"},
            "rimborsi_pagati": {"$sum": {"$cond": [{"$eq": ["$stato", "pagato"]}, 1, 0]}},
            "importo_pagato": {"$sum": {"$cond": [{"$eq": ["$stato", "pagato"]}, "$importo_totale", 0]}},
        }},
    ]

    results = []
    async for result in db.rimborsi.aggregate(pipeline):
        user_doc = await db.users.find_one({"_id": ObjectId(result["_id"])})
        if user_doc:
            result["user_nome"] = f"{user_doc['nome']} {user_doc['cognome']}"
            result["user_email"] = user_doc["email"]
            result["user_iban"] = user_doc.get("iban", "")
        results.append(result)

    return results


@router.get("/reports/rimborsi-export")
async def export_rimborsi(request: Request, anno: int, formato: str = "csv"):
    user = await get_current_user(request)

    if not user_has_any_role(user, ["admin", "cassiere", "superadmin", "superuser"]):
        raise HTTPException(status_code=403, detail="Permessi insufficienti")

    query = {"data": {"$regex": f"^{anno}"}}
    if not user_has_any_role(user, ["superadmin", "superuser"]):
        query["sede_id"] = user.get("sede_id")

    rimborsi = []
    async for rimborso in db.rimborsi.find(query).sort("data", 1):
        rimborso_user = await db.users.find_one({"_id": ObjectId(rimborso["user_id"])})
        motivo = await db.motivi_rimborso.find_one({"_id": ObjectId(rimborso["motivo_id"])}) if rimborso.get("motivo_id") else None

        rimborsi.append({
            "Data": rimborso["data"],
            "Utente": f"{rimborso_user['nome']} {rimborso_user['cognome']}" if rimborso_user else "N/A",
            "Email": rimborso_user["email"] if rimborso_user else "",
            "IBAN": rimborso_user.get("iban", "") if rimborso_user else "",
            "Motivo": motivo["nome"] if motivo else "N/A",
            "Partenza": rimborso["indirizzo_partenza"],
            "Arrivo": rimborso["indirizzo_arrivo"],
            "KM Totali": rimborso["km_totali"],
            "Importo KM": f"{rimborso['importo_km']:.2f}",
            "Importo Pasti": f"{rimborso.get('importo_pasti', 0):.2f}",
            "Autostrada": f"{rimborso.get('costo_autostrada', 0):.2f}",
            "Totale": f"{rimborso['importo_totale']:.2f}",
            "Stato": rimborso["stato"],
            "Note": rimborso.get("note", ""),
        })

    if formato == "csv":
        output = io.StringIO()
        if rimborsi:
            writer = csv.DictWriter(output, fieldnames=rimborsi[0].keys())
            writer.writeheader()
            writer.writerows(rimborsi)

        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=rimborsi_{anno}.csv"},
        )
    elif formato == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = f"Rimborsi {anno}"

        headers = list(rimborsi[0].keys()) if rimborsi else [
            "Data", "Utente", "Email", "IBAN", "Motivo", "Partenza", "Arrivo",
            "KM Totali", "Importo KM", "Importo Pasti", "Autostrada", "Totale", "Stato", "Note",
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E4D8C", end_color="1E4D8C", fill_type="solid")
        thin = Side(border_style="thin", color="CCCCCC")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row_idx, r in enumerate(rimborsi, start=2):
            for col_idx, key in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=r.get(key, ""))
                cell.border = border
                cell.alignment = Alignment(vertical="center")

        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        ws.row_dimensions[1].height = 25

        if rimborsi:
            total = sum(float(r["Totale"]) for r in rimborsi)
            total_row = len(rimborsi) + 2
            ws.cell(row=total_row, column=1, value="TOTALE").font = Font(bold=True)
            cell = ws.cell(row=total_row, column=headers.index("Totale") + 1, value=f"{total:.2f}")
            cell.font = Font(bold=True, color="1E4D8C")

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return StreamingResponse(
            iter([out.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=rimborsi_{anno}.xlsx"},
        )
    elif formato == "pdf":
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        out = io.BytesIO()
        doc = SimpleDocTemplate(
            out, pagesize=landscape(A4),
            rightMargin=10*mm, leftMargin=10*mm, topMargin=15*mm, bottomMargin=15*mm,
        )

        elements = []
        styles = getSampleStyleSheet()
        title_style = styles["Title"]
        title_style.textColor = colors.HexColor("#1E4D8C")

        sede_nome = ""
        if user.get("sede_id"):
            sede = await db.sedi.find_one({"_id": ObjectId(user["sede_id"])})
            if sede:
                sede_nome = f" - {sede['nome']}"

        elements.append(Paragraph(f"Rendiconto Rimborsi {anno}{sede_nome}", title_style))
        elements.append(Paragraph(
            f"<font size=9 color='#666'>Generato da {user.get('nome', '')} {user.get('cognome', '')} il {datetime.now().strftime('%d/%m/%Y %H:%M')}</font>",
            styles["Normal"],
        ))
        elements.append(Spacer(1, 8*mm))

        pdf_headers = ["Data", "Utente", "Motivo", "KM", "Importo €", "Stato"]
        data = [pdf_headers]
        total = 0.0
        for r in rimborsi:
            data.append([
                r["Data"],
                r["Utente"],
                r["Motivo"],
                str(r["KM Totali"]),
                r["Totale"],
                r["Stato"].upper(),
            ])
            try:
                total += float(r["Totale"])
            except (ValueError, TypeError):
                pass

        data.append(["", "", "", "TOTALE", f"{total:.2f}", ""])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E4D8C")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor("#F5F7FA")]),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#FFF7E0")),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)

        doc.build(elements)
        out.seek(0)
        return StreamingResponse(
            iter([out.getvalue()]),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=rimborsi_{anno}.pdf"},
        )
    else:
        return rimborsi
