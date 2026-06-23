"""
==============================================
SLA PORTALE - Server FastAPI
==============================================

Entry point dell'applicazione backend.
Gestisce tutte le API per il portale rimborsi del Sindacato Lavoratori Autostradali.

Funzionalità principali:
- Autenticazione JWT multi-ruolo (7 ruoli)
- Gestione rimborsi con calcolo KM automatico/manuale
- Upload documenti e modulistica
- Bacheca e notifiche
- Export PDF/Excel rendiconti

Per avviare in sviluppo:
    uvicorn server:app --reload --port 8001

Documentazione API automatica:
    http://localhost:8001/docs (Swagger)
    http://localhost:8001/redoc (ReDoc)

==============================================
"""

# ==================== IMPORTS ====================

# FastAPI framework
from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, UploadFile, File, Form
from fastapi.responses import FileResponse, StreamingResponse
from starlette.middleware.cors import CORSMiddleware

# Database - MongoDB async driver
from bson import ObjectId

# Standard library
import os
import io
import csv
import uuid
import math
import asyncio
from datetime import datetime, timezone, timedelta
from typing import List, Optional
from pathlib import Path

# External APIs
import aiofiles
import httpx
import jwt

# Data validation
from pydantic import BaseModel

# === Core moduli refactorizzati ===
from core.config import (
    UPLOAD_DIR, MAX_FILE_SIZE, GOOGLE_MAPS_API_KEY,
    FRONTEND_ORIGIN_REGEX, JWT_SECRET, JWT_ALGORITHM, logger,
)
from core.db import client, db
from core.auth import (
    hash_password, verify_password, validate_password_strength,
    generate_temporary_password, create_access_token, create_refresh_token,
    get_current_user,
)
from core.roles import (
    VALID_ROLES, _user_roles, user_has_role, user_has_any_role,
    normalize_roles_input,
)
from core.notifications import (
    _notify_users_by_role, _notify_user, _notify_all_in_sede,
)
from core.audit import _log_audit
from core.scheduler import (
    _check_pending_reimbursements, _pending_reimbursements_scheduler,
)
from models_api import (
    UserBase, UserCreate, UserUpdate, UserResponse, LoginRequest,
    ChangePasswordRequest, ToggleDisableRequest, UpdateRuoliRequest,
    TOTPEnableRequest, TOTPDisableRequest,
    SedeCreate, SedeUpdate,
    MotivoRimborsoCreate, MotivoRimborsoUpdate,
    RimborsoCreate, RimborsoUpdate,
    AnnuncioCreate, DocumentoCreate, CalcoloKmRequest,
    ContattoBase, ContattoCreate, ContattoUpdate,
)

# ==================== CONFIGURAZIONE APP ====================

# FastAPI App
APP_VERSION = "0.9.3-beta"
APP_BUILD_DATE = "2026-02-15"
APP_RELEASE_NAME = "Fix healthcheck frontend (IPv4)"

app = FastAPI(
    title="SLA Sindacato - Portale Rimborsi",
    description="Gestione rimborsi e documenti per 30 concessionarie autostradali",
    version=APP_VERSION,
)
api_router = APIRouter(prefix="/api")


@api_router.get("/version")
async def get_version():
    """
    GET /api/version
    Restituisce versione applicazione (pubblico, no auth).
    Utile per verificare che frontend e backend siano allineati dopo deploy.
    """
    return {
        "version": APP_VERSION,
        "build_date": APP_BUILD_DATE,
        "release_name": APP_RELEASE_NAME,
    }

# ==================== AUTH ROUTES ====================
# API per registrazione, login, logout e gestione sessioni

@api_router.post("/auth/register")
async def register(user_data: UserCreate, response: Response):
    """
    POST /api/auth/register
    Registrazione nuovo utente
    
    Regole:
    - Email univoca
    - Auto-registrazione solo per "iscritto" o "delegato"
    - "Iscritto": IBAN e indirizzo NON richiesti
    - "Delegato": IBAN e indirizzo OBBLIGATORI (per ricevere rimborsi)
    - Altri ruoli possono essere assegnati solo da admin
    """
    email = user_data.email.lower()
    
    # Valida la password (8 char, lettera+numero+speciale)
    validate_password_strength(user_data.password)
    
    # Check email già registrata
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=400, detail="Email già registrata")
    
    # Only allow iscritto or delegato for self-registration
    allowed_roles = ["iscritto", "delegato"]
    ruolo = user_data.ruolo if user_data.ruolo in allowed_roles else "iscritto"
    
    # Validate sede exists if provided
    if user_data.sede_id:
        sede = await db.sedi.find_one({"_id": ObjectId(user_data.sede_id)})
        if not sede:
            raise HTTPException(status_code=400, detail="Sede non trovata")
    
    # REGOLA: Delegato DEVE avere IBAN e indirizzo (per rimborsi)
    if ruolo == "delegato":
        if not user_data.indirizzo:
            raise HTTPException(status_code=400, detail="Indirizzo obbligatorio per i delegati")
        if not user_data.iban:
            raise HTTPException(status_code=400, detail="IBAN obbligatorio per i delegati")
    
    user_doc = {
        "email": email,
        "password_hash": hash_password(user_data.password),
        "nome": user_data.nome,
        "cognome": user_data.cognome,
        "telefono": user_data.telefono,
        "indirizzo": user_data.indirizzo if ruolo != "iscritto" else None,
        "citta": user_data.citta if ruolo != "iscritto" else None,
        "cap": user_data.cap if ruolo != "iscritto" else None,
        "iban": user_data.iban if ruolo != "iscritto" else None,
        "ruolo": ruolo,
        "ruoli": [ruolo],
        "sede_id": user_data.sede_id,
        "disabled": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    result = await db.users.insert_one(user_doc)
    user_id = str(result.inserted_id)
    
    access_token = create_access_token(user_id, email)
    refresh_token = create_refresh_token(user_id)
    
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=True, samesite="none", max_age=86400, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=True, samesite="none", max_age=604800, path="/")
    
    user_doc["id"] = user_id
    user_doc.pop("password_hash")
    user_doc.pop("_id", None)
    user_doc["ruoli"] = user_doc.get("ruoli") or [user_doc.get("ruolo")]
    
    return user_doc

@api_router.post("/auth/login")
async def login(login_data: LoginRequest, request: Request, response: Response):
    email = login_data.email.lower()
    identifier = f"{request.client.host}:{email}"
    
    attempts = await db.login_attempts.find_one({"identifier": identifier})
    if attempts and attempts.get("count", 0) >= 5:
        lockout_time = attempts.get("last_attempt")
        if lockout_time:
            lockout_dt = datetime.fromisoformat(lockout_time) if isinstance(lockout_time, str) else lockout_time
            if datetime.now(timezone.utc) - lockout_dt < timedelta(minutes=15):
                raise HTTPException(status_code=429, detail="Troppi tentativi. Riprova tra 15 minuti.")
    
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(login_data.password, user["password_hash"]):
        await db.login_attempts.update_one(
            {"identifier": identifier},
            {"$inc": {"count": 1}, "$set": {"last_attempt": datetime.now(timezone.utc).isoformat()}},
            upsert=True
        )
        raise HTTPException(status_code=401, detail="Credenziali non valide")
    
    # Blocca login se utente disattivato
    if user.get("disabled", False):
        raise HTTPException(status_code=403, detail="Account disattivato. Contatta l'amministratore.")
    
    # === 2FA TOTP (Two-Factor Authentication) ===
    # Se l'utente ha 2FA attivo, richiedi il codice prima di completare il login.
    if user.get("totp_enabled"):
        from core.totp import verify_code
        if not login_data.totp_code:
            # 2FA richiesto ma non fornito → il client deve mostrare campo "Codice 2FA"
            raise HTTPException(status_code=401, detail="2FA_REQUIRED")
        if not verify_code(user.get("totp_secret", ""), login_data.totp_code):
            await db.login_attempts.update_one(
                {"identifier": identifier},
                {"$inc": {"count": 1}, "$set": {"last_attempt": datetime.now(timezone.utc).isoformat()}},
                upsert=True
            )
            raise HTTPException(status_code=401, detail="Codice 2FA non valido")
    
    await db.login_attempts.delete_one({"identifier": identifier})
    
    user_id = str(user["_id"])
    access_token = create_access_token(user_id, email)
    refresh_token = create_refresh_token(user_id)
    
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=True, samesite="none", max_age=86400, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=True, samesite="none", max_age=604800, path="/")
    
    user_response = {
        "id": user_id,
        "email": user["email"],
        "nome": user["nome"],
        "cognome": user["cognome"],
        "telefono": user.get("telefono"),
        "indirizzo": user.get("indirizzo"),
        "citta": user.get("citta"),
        "cap": user.get("cap"),
        "iban": user.get("iban"),
        "ruolo": user["ruolo"],
        "ruoli": user.get("ruoli") or [user["ruolo"]],
        "sede_id": user.get("sede_id"),
        "totp_enabled": bool(user.get("totp_enabled")),
        "created_at": user["created_at"]
    }
    
    if user.get("sede_id"):
        sede = await db.sedi.find_one({"_id": ObjectId(user["sede_id"])})
        if sede:
            user_response["sede_nome"] = sede["nome"]
    
    return user_response

@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie(key="access_token", path="/")
    response.delete_cookie(key="refresh_token", path="/")
    return {"message": "Logout effettuato"}

@api_router.get("/auth/me")
async def get_me(request: Request):
    user = await get_current_user(request)
    # Espone solo lo stato 2FA, non il segreto
    user["totp_enabled"] = bool(user.get("totp_enabled"))
    user.pop("totp_secret", None)
    return user


# ==================== 2FA TOTP ENDPOINTS ====================
# Disponibile solo per ruoli sensibili (admin, superadmin).
# Gli iscritti/delegati non hanno 2FA (UX semplice per l'utenza ampia).

def _2fa_allowed(user: dict) -> bool:
    return user_has_any_role(user, ["admin", "superadmin"])


@api_router.post("/auth/2fa/setup")
async def setup_2fa(request: Request):
    """
    Inizia la configurazione 2FA: genera segreto + QR code.
    Il segreto viene salvato come 'pending' finché non viene verificato il primo codice.
    """
    from core.totp import generate_secret, generate_qrcode_png
    user = await get_current_user(request)
    if not _2fa_allowed(user):
        raise HTTPException(status_code=403, detail="2FA disponibile solo per admin/superadmin")
    if user.get("totp_enabled"):
        raise HTTPException(status_code=400, detail="2FA già attivo. Disabilitalo prima di riconfigurarlo.")
    
    secret = generate_secret()
    qr = generate_qrcode_png(secret, user["email"])
    
    # Salva il segreto come "pending" (NON ancora abilitato)
    await db.users.update_one(
        {"_id": ObjectId(user["id"])},
        {"$set": {"totp_secret_pending": secret}}
    )
    
    return {
        "secret": secret,  # Solo per debug/manual entry sull'app authenticator
        "qrcode": qr,
        "issuer": "Portale SLA",
        "account": user["email"],
    }


@api_router.post("/auth/2fa/enable")
async def enable_2fa(payload: TOTPEnableRequest, request: Request):
    """
    Verifica il primo codice e attiva definitivamente il 2FA.
    """
    from core.totp import verify_code
    user = await get_current_user(request)
    if not _2fa_allowed(user):
        raise HTTPException(status_code=403, detail="2FA disponibile solo per admin/superadmin")
    
    user_doc = await db.users.find_one({"_id": ObjectId(user["id"])})
    pending = user_doc.get("totp_secret_pending")
    if not pending:
        raise HTTPException(status_code=400, detail="Nessuna configurazione 2FA in corso. Riavvia da Setup.")
    
    if not verify_code(pending, payload.code):
        raise HTTPException(status_code=400, detail="Codice non valido")
    
    await db.users.update_one(
        {"_id": ObjectId(user["id"])},
        {
            "$set": {"totp_secret": pending, "totp_enabled": True},
            "$unset": {"totp_secret_pending": ""}
        }
    )
    await _log_audit(
        actor=user,
        action="user.enable_2fa",
        target_type="user",
        target_id=user["id"],
        target_label=user["email"],
    )
    return {"message": "2FA attivato con successo", "totp_enabled": True}


@api_router.post("/auth/2fa/disable")
async def disable_2fa(payload: TOTPDisableRequest, request: Request):
    """
    Disabilita 2FA. Richiede la password per evitare disabilitazioni accidentali
    (es. se qualcuno ti ruba il cookie e non sa la password, non può togliere 2FA).
    """
    user = await get_current_user(request)
    user_doc = await db.users.find_one({"_id": ObjectId(user["id"])})
    
    if not verify_password(payload.password, user_doc["password_hash"]):
        raise HTTPException(status_code=401, detail="Password non valida")
    
    if not user_doc.get("totp_enabled"):
        raise HTTPException(status_code=400, detail="2FA non è attivo")
    
    await db.users.update_one(
        {"_id": ObjectId(user["id"])},
        {"$unset": {"totp_secret": "", "totp_enabled": "", "totp_secret_pending": ""}}
    )
    await _log_audit(
        actor=user,
        action="user.disable_2fa",
        target_type="user",
        target_id=user["id"],
        target_label=user["email"],
    )
    return {"message": "2FA disattivato", "totp_enabled": False}

@api_router.post("/auth/refresh")
async def refresh_token(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="Token di refresh non trovato")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Token non valido")
        user_id = payload["sub"]
        user = await db.users.find_one({"_id": ObjectId(user_id)})
        if not user:
            raise HTTPException(status_code=401, detail="Utente non trovato")
        access_token = create_access_token(user_id, user["email"])
        response.set_cookie(key="access_token", value=access_token, httponly=True, secure=True, samesite="none", max_age=86400, path="/")
        return {"message": "Token aggiornato"}
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token scaduto")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token non valido")


# ==================== PASSWORD CHANGE / RESET ====================


@api_router.post("/auth/change-password")
async def change_password(data: ChangePasswordRequest, request: Request, response: Response):
    """
    Utente loggato cambia la propria password.
    Richiede password attuale. Forza re-login dopo il cambio.
    """
    user = await get_current_user(request)
    
    user_doc = await db.users.find_one({"_id": ObjectId(user["id"])})
    if not user_doc:
        raise HTTPException(status_code=404, detail="Utente non trovato")
    
    if not verify_password(data.current_password, user_doc["password_hash"]):
        raise HTTPException(status_code=400, detail="Password attuale non corretta")
    
    if data.current_password == data.new_password:
        raise HTTPException(status_code=400, detail="La nuova password deve essere diversa da quella attuale")
    
    validate_password_strength(data.new_password)
    
    await db.users.update_one(
        {"_id": ObjectId(user["id"])},
        {"$set": {
            "password_hash": hash_password(data.new_password),
            "password_changed_at": datetime.now(timezone.utc).isoformat(),
            "must_change_password": False
        }}
    )
    
    # Forza re-login
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    
    return {"message": "Password aggiornata. Effettua nuovamente il login per sicurezza."}


@api_router.post("/users/{user_id}/reset-password")
async def admin_reset_password(user_id: str, request: Request):
    """
    Admin/Segretario/SuperAdmin genera password temporanea per un utente.
    Restituisce la password in chiaro (mostrata UNA SOLA VOLTA).
    """
    current_user = await get_current_user(request)
    
    if not user_has_any_role(current_user, ["admin", "segretario", "superadmin"]):
        raise HTTPException(status_code=403, detail="Permessi insufficienti")
    
    target_user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not target_user:
        raise HTTPException(status_code=404, detail="Utente non trovato")
    
    if not user_has_role(current_user, "superadmin"):
        if target_user.get("sede_id") != current_user.get("sede_id"):
            raise HTTPException(status_code=403, detail="Non autorizzato per questa sede")
        if user_has_any_role(target_user, ["superadmin", "superuser"]):
            raise HTTPException(status_code=403, detail="Non autorizzato a resettare questo ruolo")
    
    temp_password = generate_temporary_password(12)
    
    await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": {
            "password_hash": hash_password(temp_password),
            "password_reset_at": datetime.now(timezone.utc).isoformat(),
            "password_reset_by": current_user["id"],
            "must_change_password": True
        }}
    )
    
    # Notifica all'utente
    await db.notifiche.insert_one({
        "user_id": user_id,
        "sede_id": target_user.get("sede_id"),
        "tipo": "sicurezza",
        "titolo": "Password reimpostata",
        "messaggio": f"La tua password è stata reimpostata da {current_user['nome']} {current_user['cognome']}. Accedi con la nuova password ricevuta e cambiala al primo accesso.",
        "letto": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    await _log_audit(
        actor=current_user,
        action="user.reset_password",
        target_type="user",
        target_id=user_id,
        target_label=f"{target_user.get('nome', '')} {target_user.get('cognome', '')} ({target_user.get('email', '')})".strip(),
        sede_id=target_user.get("sede_id"),
    )
    
    return {
        "message": "Password reimpostata con successo",
        "temporary_password": temp_password,
        "user_email": target_user["email"]
    }


# ==================== DISATTIVA / CANCELLA UTENTE ====================


@api_router.put("/users/{user_id}/toggle-disabled")
async def toggle_user_disabled(user_id: str, data: ToggleDisableRequest, request: Request):
    """
    Admin/Segretario/SuperAdmin disattiva o riattiva un utente (soft).
    Utente disattivato non può loggarsi ma i dati storici restano intatti.
    """
    current_user = await get_current_user(request)
    
    if not user_has_any_role(current_user, ["admin", "segretario", "superadmin"]):
        raise HTTPException(status_code=403, detail="Permessi insufficienti")
    
    if user_id == current_user["id"]:
        raise HTTPException(status_code=400, detail="Non puoi disattivare il tuo stesso account")
    
    target_user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not target_user:
        raise HTTPException(status_code=404, detail="Utente non trovato")
    
    if not user_has_role(current_user, "superadmin"):
        if target_user.get("sede_id") != current_user.get("sede_id"):
            raise HTTPException(status_code=403, detail="Non autorizzato per questa sede")
        if user_has_any_role(target_user, ["superadmin", "superuser"]):
            raise HTTPException(status_code=403, detail="Non autorizzato a modificare questo ruolo")
    
    update_fields = {
        "disabled": data.disabled,
        "disabled_at": datetime.now(timezone.utc).isoformat() if data.disabled else None,
        "disabled_by": current_user["id"] if data.disabled else None,
    }
    
    await db.users.update_one({"_id": ObjectId(user_id)}, {"$set": update_fields})
    
    await _log_audit(
        actor=current_user,
        action="user.disable" if data.disabled else "user.enable",
        target_type="user",
        target_id=user_id,
        target_label=f"{target_user.get('nome', '')} {target_user.get('cognome', '')} ({target_user.get('email', '')})".strip(),
        sede_id=target_user.get("sede_id"),
        old_value="active" if not target_user.get("disabled") else "disabled",
        new_value="disabled" if data.disabled else "active",
    )
    
    return {
        "message": "Utente disattivato" if data.disabled else "Utente riattivato",
        "disabled": data.disabled
    }


@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, request: Request):
    """
    Cancellazione definitiva utente. Solo SuperAdmin.
    Rimborsi/annunci/documenti restano ma con riferimento "[utente eliminato]".
    """
    current_user = await get_current_user(request)
    
    if not user_has_role(current_user, "superadmin"):
        raise HTTPException(status_code=403, detail="Solo SuperAdmin può cancellare definitivamente gli utenti")
    
    if user_id == current_user["id"]:
        raise HTTPException(status_code=400, detail="Non puoi cancellare il tuo stesso account")
    
    target_user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not target_user:
        raise HTTPException(status_code=404, detail="Utente non trovato")
    
    user_label = "[utente eliminato]"
    
    # Anonimizza riferimenti nei rimborsi (manteniamo storico)
    await db.rimborsi.update_many(
        {"user_id": user_id},
        {"$set": {"user_nome": user_label, "user_eliminato": True}}
    )
    
    # Anonimizza autore in annunci
    await db.annunci.update_many(
        {"autore_id": user_id},
        {"$set": {"autore_nome": user_label, "autore_eliminato": True}}
    )
    
    # Cancella notifiche dirette all'utente
    await db.notifiche.delete_many({"user_id": user_id})
    
    # Cancella utente
    await db.users.delete_one({"_id": ObjectId(user_id)})
    
    await _log_audit(
        actor=current_user,
        action="user.delete",
        target_type="user",
        target_id=user_id,
        target_label=f"{target_user.get('nome', '')} {target_user.get('cognome', '')} ({target_user.get('email', '')})".strip(),
        sede_id=target_user.get("sede_id"),
        old_value=target_user.get("ruolo"),
        note="Cancellazione definitiva, dati anonimizzati nei record storici",
    )
    
    return {"message": "Utente cancellato definitivamente"}



# ==================== AUDIT LOG ROUTES ====================

@api_router.get("/audit-log")
async def get_audit_log(
    request: Request,
    target_type: Optional[str] = None,
    target_id: Optional[str] = None,
    action: Optional[str] = None,
    limit: int = 100,
):
    """
    Restituisce gli eventi audit.
    - SuperAdmin/SuperUser: vede tutto
    - Admin/Cassiere/Segretario: vede solo la propria sede
    - Altri ruoli: forbidden
    
    Filtri opzionali: target_type, target_id, action
    """
    user = await get_current_user(request)
    
    if not user_has_any_role(user, ["superadmin", "superuser", "admin", "cassiere", "segretario"]):
        raise HTTPException(status_code=403, detail="Permessi insufficienti")
    
    query = {}
    if not user_has_any_role(user, ["superadmin", "superuser"]):
        query["sede_id"] = user.get("sede_id")
    
    if target_type:
        query["target_type"] = target_type
    if target_id:
        query["target_id"] = target_id
    if action:
        query["action"] = action
    
    limit = min(max(limit, 1), 500)
    
    entries = []
    async for entry in db.audit_log.find(query).sort("created_at", -1).limit(limit):
        entry["id"] = str(entry["_id"])
        entry.pop("_id")
        entries.append(entry)
    
    return entries



# ==================== GOOGLE MAPS ====================

@api_router.post("/calcola-km")
async def calcola_km(data: CalcoloKmRequest, request: Request):
    await get_current_user(request)
    
    if not GOOGLE_MAPS_API_KEY:
        raise HTTPException(status_code=500, detail="Google Maps API non configurata")
    
    try:
        async with httpx.AsyncClient() as client:
            response = await client.get(
                "https://maps.googleapis.com/maps/api/directions/json",
                params={
                    "origin": data.origine,
                    "destination": data.destinazione,
                    "key": GOOGLE_MAPS_API_KEY,
                    "language": "it"
                }
            )
            result = response.json()
            
            if result.get("status") != "OK":
                raise HTTPException(status_code=400, detail=f"Impossibile calcolare il percorso: {result.get('status')}")
            
            # Get distance in meters and convert to km, round up
            distance_meters = result["routes"][0]["legs"][0]["distance"]["value"]
            distance_km = math.ceil(distance_meters / 1000)
            
            return {
                "km": distance_km,
                "origine": result["routes"][0]["legs"][0]["start_address"],
                "destinazione": result["routes"][0]["legs"][0]["end_address"],
                "durata": result["routes"][0]["legs"][0]["duration"]["text"]
            }
    except httpx.RequestError as e:
        raise HTTPException(status_code=500, detail=f"Errore di connessione: {str(e)}")

# ==================== SEDI ROUTES ====================

@api_router.get("/sedi")
async def get_sedi(request: Request):
    # Allow unauthenticated access for registration
    sedi = []
    async for sede in db.sedi.find({}, {"_id": 1, "nome": 1, "codice": 1, "indirizzo": 1, "tariffa_km": 1, "rimborso_pasti": 1, "rimborso_autostrada": 1}):
        sede["id"] = str(sede["_id"])
        sede.pop("_id")
        sedi.append(sede)
    return sedi

@api_router.post("/sedi")
async def create_sede(sede_data: SedeCreate, request: Request):
    user = await get_current_user(request)
    if not user_has_any_role(user, ["superadmin"]):
        raise HTTPException(status_code=403, detail="Solo il SuperAdmin può creare sedi")
    
    existing = await db.sedi.find_one({"codice": sede_data.codice})
    if existing:
        raise HTTPException(status_code=400, detail="Codice sede già esistente")
    
    sede_doc = {
        "nome": sede_data.nome,
        "codice": sede_data.codice,
        "indirizzo": sede_data.indirizzo,
        "tariffa_km": sede_data.tariffa_km,
        "rimborso_pasti": sede_data.rimborso_pasti,
        "rimborso_autostrada": sede_data.rimborso_autostrada,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    result = await db.sedi.insert_one(sede_doc)
    sede_doc["id"] = str(result.inserted_id)
    sede_doc.pop("_id", None)
    return sede_doc

@api_router.put("/sedi/{sede_id}")
async def update_sede(sede_id: str, sede_data: SedeUpdate, request: Request):
    user = await get_current_user(request)
    if not user_has_any_role(user, ["superadmin", "admin"]):
        raise HTTPException(status_code=403, detail="Permessi insufficienti")
    
    update_data = {k: v for k, v in sede_data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Nessun dato da aggiornare")
    
    result = await db.sedi.update_one({"_id": ObjectId(sede_id)}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Sede non trovata")
    
    return {"message": "Sede aggiornata"}

@api_router.delete("/sedi/{sede_id}")
async def delete_sede(sede_id: str, request: Request):
    user = await get_current_user(request)
    if not user_has_any_role(user, ["superadmin"]):
        raise HTTPException(status_code=403, detail="Solo il SuperAdmin può eliminare sedi")
    
    result = await db.sedi.delete_one({"_id": ObjectId(sede_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Sede non trovata")
    
    return {"message": "Sede eliminata"}

# ==================== MOTIVI RIMBORSO ROUTES ====================

@api_router.get("/motivi-rimborso")
async def get_motivi_rimborso(request: Request):
    await get_current_user(request)
    motivi = []
    async for motivo in db.motivi_rimborso.find({}):
        motivo["id"] = str(motivo["_id"])
        motivo.pop("_id")
        motivi.append(motivo)
    return motivi

@api_router.post("/motivi-rimborso")
async def create_motivo_rimborso(motivo_data: MotivoRimborsoCreate, request: Request):
    user = await get_current_user(request)
    if not user_has_any_role(user, ["superadmin"]):
        raise HTTPException(status_code=403, detail="Solo il SuperAdmin può gestire i motivi")
    
    motivo_doc = {
        "nome": motivo_data.nome,
        "descrizione": motivo_data.descrizione,
        "richiede_note": motivo_data.richiede_note,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    result = await db.motivi_rimborso.insert_one(motivo_doc)
    motivo_doc["id"] = str(result.inserted_id)
    motivo_doc.pop("_id", None)
    return motivo_doc

@api_router.put("/motivi-rimborso/{motivo_id}")
async def update_motivo_rimborso(motivo_id: str, motivo_data: MotivoRimborsoUpdate, request: Request):
    user = await get_current_user(request)
    if not user_has_any_role(user, ["superadmin"]):
        raise HTTPException(status_code=403, detail="Solo il SuperAdmin può gestire i motivi")
    
    update_data = {k: v for k, v in motivo_data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Nessun dato da aggiornare")
    
    result = await db.motivi_rimborso.update_one({"_id": ObjectId(motivo_id)}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Motivo non trovato")
    
    return {"message": "Motivo aggiornato"}

@api_router.delete("/motivi-rimborso/{motivo_id}")
async def delete_motivo_rimborso(motivo_id: str, request: Request):
    user = await get_current_user(request)
    if not user_has_any_role(user, ["superadmin"]):
        raise HTTPException(status_code=403, detail="Solo il SuperAdmin può gestire i motivi")
    
    result = await db.motivi_rimborso.delete_one({"_id": ObjectId(motivo_id)})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Motivo non trovato")
    
    return {"message": "Motivo eliminato"}

# ==================== RIMBORSI ROUTES ====================

@api_router.get("/rimborsi")
async def get_rimborsi(
    request: Request,
    stato: Optional[str] = None,
    anno: Optional[int] = None,
    data_da: Optional[str] = None,        # YYYY-MM-DD
    data_a: Optional[str] = None,         # YYYY-MM-DD
    user_id: Optional[str] = None,        # filtro per autore
    sede_id: Optional[str] = None,        # filtro per sede (solo super*)
    motivo_id: Optional[str] = None,      # filtro per motivo
    importo_min: Optional[float] = None,  # range importo
    importo_max: Optional[float] = None,
):
    user = await get_current_user(request)
    
    query: dict = {}
    
    # Scope: superadmin/superuser → tutti; admin/cassiere → sede; altri → solo i propri
    if not user_has_any_role(user, ["superadmin", "superuser"]):
        if user_has_any_role(user, ["admin", "cassiere"]):
            query["sede_id"] = user.get("sede_id")
        else:
            query["user_id"] = user["id"]
    
    if stato:
        query["stato"] = stato
    
    # Filtro per anno (legacy) o range date
    if data_da or data_a:
        date_range: dict = {}
        if data_da:
            date_range["$gte"] = data_da
        if data_a:
            date_range["$lte"] = data_a
        query["data"] = date_range
    elif anno:
        query["data"] = {"$regex": f"^{anno}"}
    
    # Filtro per autore (solo per chi può vedere multi-utente)
    if user_id and user_has_any_role(user, ["admin", "cassiere", "superadmin", "superuser"]):
        query["user_id"] = user_id
    
    # Filtro sede - solo super* può filtrare diverse sedi
    if sede_id and user_has_any_role(user, ["superadmin", "superuser"]):
        query["sede_id"] = sede_id
    
    if motivo_id:
        query["motivo_id"] = motivo_id
    
    # Range importo
    if importo_min is not None or importo_max is not None:
        importo_range: dict = {}
        if importo_min is not None:
            importo_range["$gte"] = importo_min
        if importo_max is not None:
            importo_range["$lte"] = importo_max
        query["importo_totale"] = importo_range
    
    rimborsi = []
    async for rimborso in db.rimborsi.find(query).sort("created_at", -1):
        rimborso["id"] = str(rimborso["_id"])
        rimborso.pop("_id")
        
        rimborso_user = await db.users.find_one({"_id": ObjectId(rimborso["user_id"])})
        if rimborso_user:
            rimborso["user_nome"] = f"{rimborso_user['nome']} {rimborso_user['cognome']}"
        
        if rimborso.get("motivo_id"):
            motivo = await db.motivi_rimborso.find_one({"_id": ObjectId(rimborso["motivo_id"])})
            if motivo:
                rimborso["motivo_nome"] = motivo["nome"]
        
        rimborsi.append(rimborso)
    
    return rimborsi

@api_router.post("/rimborsi")
async def create_rimborso(rimborso_data: RimborsoCreate, request: Request):
    """
    POST /api/rimborsi
    Crea una nuova richiesta di rimborso
    
    Flusso:
    1. Verifica permessi (ruolo iscritto NON può creare rimborsi)
    2. Verifica note obbligatorie se motivo="Altro"
    3. Calcola totale: (KM * tariffa) + pasti + autostrada
    4. Crea notifica per admin se KM modificati manualmente
    
    Regole speciali:
    - Se km_modificati_manualmente=True, genera alert per admin
    - Ricevute pasti: nessun limite di costo (campo importo_pasti libero)
    """
    user = await get_current_user(request)
    
    # REGOLA: Solo ruoli con accesso alla sezione rimborsi possono crearli
    if not user_has_any_role(user, ["delegato", "segreteria", "segretario", "admin", "superadmin"]):
        raise HTTPException(status_code=403, detail="Gli iscritti non possono richiedere rimborsi")
    
    # Check if motivo requires note (es: "Altro" richiede sempre note)
    motivo = await db.motivi_rimborso.find_one({"_id": ObjectId(rimborso_data.motivo_id)})
    if motivo and motivo.get("richiede_note") and not rimborso_data.note:
        raise HTTPException(status_code=400, detail="Per questo motivo le note sono obbligatorie")
    
    # Get sede tariffs (tariffa chilometrica della sede)
    sede = None
    if user.get("sede_id"):
        sede = await db.sedi.find_one({"_id": ObjectId(user["sede_id"])})
    
    tariffa_km = sede["tariffa_km"] if sede else 0.35
    
    # CALCOLO TOTALE RIMBORSO
    km_totali = rimborso_data.km_andata * (2 if rimborso_data.andata_ritorno else 1)
    importo_km = km_totali * tariffa_km
    importo_autostrada = rimborso_data.costo_autostrada if rimborso_data.uso_autostrada else 0
    importo_pasti = rimborso_data.importo_pasti  # Nessun limite, inserito dall'utente
    importo_totale = importo_km + importo_pasti + importo_autostrada
    
    rimborso_doc = {
        "user_id": user["id"],
        "sede_id": user.get("sede_id"),
        "data": rimborso_data.data,
        "motivo_id": rimborso_data.motivo_id,
        "indirizzo_partenza": rimborso_data.indirizzo_partenza,
        "indirizzo_partenza_tipo": rimborso_data.indirizzo_partenza_tipo,
        "indirizzo_arrivo": rimborso_data.indirizzo_arrivo,
        "km_andata": rimborso_data.km_andata,
        "km_calcolati": rimborso_data.km_calcolati,
        "km_modificati_manualmente": rimborso_data.km_modificati_manualmente,
        "andata_ritorno": rimborso_data.andata_ritorno,
        "km_totali": km_totali,
        "uso_autostrada": rimborso_data.uso_autostrada,
        "costo_autostrada": importo_autostrada,
        "importo_pasti": importo_pasti,
        "numero_partecipanti_pasto": rimborso_data.numero_partecipanti_pasto,
        "tariffa_km": tariffa_km,
        "importo_km": importo_km,
        "importo_totale": importo_totale,
        "note": rimborso_data.note,
        "stato": "in_attesa",
        "ricevute": [],
        "ricevute_spese": [],
        "km_approvati": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    result = await db.rimborsi.insert_one(rimborso_doc)
    rimborso_doc["id"] = str(result.inserted_id)
    rimborso_doc.pop("_id", None)
    
    # Notifica Admin + Cassiere della sede
    notifica_msg = f"{user['nome']} {user['cognome']} ha inviato una richiesta di rimborso di €{importo_totale:.2f}"
    if rimborso_data.km_modificati_manualmente:
        notifica_msg += " - ATTENZIONE: KM modificati manualmente!"
    
    await _notify_users_by_role(
        roles=["admin", "cassiere"],
        sede_id=user.get("sede_id"),
        notifica_data={
            "tipo": "rimborso",
            "titolo": "Nuova richiesta rimborso" + (" ⚠️ KM MODIFICATI" if rimborso_data.km_modificati_manualmente else ""),
            "messaggio": notifica_msg,
            "rimborso_id": str(result.inserted_id),
            "alert_km": rimborso_data.km_modificati_manualmente,
        },
        include_global=False
    )
    
    return rimborso_doc

@api_router.post("/rimborsi/{rimborso_id}/ricevute")
async def upload_ricevuta(rimborso_id: str, request: Request, file: UploadFile = File(...)):
    user = await get_current_user(request)
    
    rimborso = await db.rimborsi.find_one({"_id": ObjectId(rimborso_id)})
    if not rimborso:
        raise HTTPException(status_code=404, detail="Rimborso non trovato")
    
    if rimborso["user_id"] != user["id"] and not user_has_any_role(user, ["admin", "superadmin"]):
        raise HTTPException(status_code=403, detail="Non autorizzato")
    
    if file.content_type not in ["application/pdf", "image/jpeg", "image/png", "image/jpg"]:
        raise HTTPException(status_code=400, detail="Formato file non supportato. Usa PDF, JPG o PNG")
    
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File troppo grande. Max 5MB")
    
    file_id = str(uuid.uuid4())
    ext = file.filename.split(".")[-1] if "." in file.filename else "pdf"
    filename = f"{file_id}.{ext}"
    filepath = UPLOAD_DIR / filename
    
    async with aiofiles.open(filepath, "wb") as f:
        await f.write(content)
    
    ricevuta = {
        "id": file_id,
        "filename": file.filename,
        "path": filename,
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.rimborsi.update_one(
        {"_id": ObjectId(rimborso_id)},
        {"$push": {"ricevute": ricevuta}}
    )
    
    return ricevuta

@api_router.post("/rimborsi/{rimborso_id}/ricevute-multi")
async def upload_ricevute_multi(rimborso_id: str, request: Request, files: List[UploadFile] = File(...)):
    """Upload multiplo di ricevute generiche al rimborso"""
    user = await get_current_user(request)
    
    rimborso = await db.rimborsi.find_one({"_id": ObjectId(rimborso_id)})
    if not rimborso:
        raise HTTPException(status_code=404, detail="Rimborso non trovato")
    
    if rimborso["user_id"] != user["id"] and not user_has_any_role(user, ["admin", "cassiere", "superadmin"]):
        raise HTTPException(status_code=403, detail="Non autorizzato")
    
    uploaded = []
    for f in files:
        if f.content_type not in ["application/pdf", "image/jpeg", "image/png", "image/jpg"]:
            continue  # Skip silently formati non supportati
        
        content = await f.read()
        if len(content) > 5 * 1024 * 1024:
            continue  # Skip files troppo grandi
        
        file_id = str(uuid.uuid4())
        ext = f.filename.split(".")[-1] if "." in f.filename else "pdf"
        filename = f"{file_id}.{ext}"
        filepath = UPLOAD_DIR / filename
        
        async with aiofiles.open(filepath, "wb") as fout:
            await fout.write(content)
        
        ricevuta = {
            "id": file_id,
            "filename": f.filename,
            "path": filename,
            "content_type": f.content_type,
            "uploaded_at": datetime.now(timezone.utc).isoformat()
        }
        uploaded.append(ricevuta)
    
    if uploaded:
        await db.rimborsi.update_one(
            {"_id": ObjectId(rimborso_id)},
            {"$push": {"ricevute": {"$each": uploaded}}}
        )
    
    return {"uploaded": uploaded, "count": len(uploaded), "skipped": len(files) - len(uploaded)}


@api_router.get("/rimborsi/{rimborso_id}/ricevute/{file_id}")
async def download_ricevuta(rimborso_id: str, file_id: str, request: Request):
    """Download di una ricevuta specifica (anche per anteprima inline)"""
    user = await get_current_user(request)
    
    rimborso = await db.rimborsi.find_one({"_id": ObjectId(rimborso_id)})
    if not rimborso:
        raise HTTPException(status_code=404, detail="Rimborso non trovato")
    
    # Permessi: proprietario o admin/cassiere/segretario
    if rimborso["user_id"] != user["id"]:
        if not user_has_any_role(user, ["admin", "cassiere", "segretario", "superadmin", "superuser"]):
            raise HTTPException(status_code=403, detail="Non autorizzato")
        if not user_has_any_role(user, ["superadmin", "superuser"]) and rimborso.get("sede_id") != user.get("sede_id"):
            raise HTTPException(status_code=403, detail="Non autorizzato per questa sede")
    
    # Cerca in ricevute e ricevute_spese
    ricevuta = next(
        (r for r in (rimborso.get("ricevute") or []) + (rimborso.get("ricevute_spese") or []) if r.get("id") == file_id),
        None
    )
    if not ricevuta:
        raise HTTPException(status_code=404, detail="Ricevuta non trovata")
    
    filepath = UPLOAD_DIR / ricevuta["path"]
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="File non trovato sul server")
    
    return FileResponse(filepath, filename=ricevuta.get("filename", "ricevuta"))


@api_router.delete("/rimborsi/{rimborso_id}/ricevute/{file_id}")
async def delete_ricevuta(rimborso_id: str, file_id: str, request: Request):
    """Elimina una ricevuta. Possibile solo se rimborso ancora 'in_attesa'."""
    user = await get_current_user(request)
    
    rimborso = await db.rimborsi.find_one({"_id": ObjectId(rimborso_id)})
    if not rimborso:
        raise HTTPException(status_code=404, detail="Rimborso non trovato")
    
    if rimborso["user_id"] != user["id"] and not user_has_any_role(user, ["admin", "cassiere", "superadmin"]):
        raise HTTPException(status_code=403, detail="Non autorizzato")
    
    if rimborso.get("stato") not in ["in_attesa", None]:
        raise HTTPException(status_code=400, detail="Impossibile rimuovere ricevute da rimborsi già approvati/pagati")
    
    # Cerca in entrambe le liste
    ricevuta = next(
        (r for r in (rimborso.get("ricevute") or []) + (rimborso.get("ricevute_spese") or []) if r.get("id") == file_id),
        None
    )
    if not ricevuta:
        raise HTTPException(status_code=404, detail="Ricevuta non trovata")
    
    # Rimuovi dal disco
    filepath = UPLOAD_DIR / ricevuta["path"]
    if filepath.exists():
        filepath.unlink()
    
    # Rimuovi da entrambe le liste
    await db.rimborsi.update_one(
        {"_id": ObjectId(rimborso_id)},
        {"$pull": {
            "ricevute": {"id": file_id},
            "ricevute_spese": {"id": file_id}
        }}
    )
    
    return {"message": "Ricevuta eliminata"}



@api_router.post("/rimborsi/{rimborso_id}/ricevute-spese")
async def upload_ricevuta_spesa(rimborso_id: str, request: Request, file: UploadFile = File(...), tipo: str = Form(...), descrizione: str = Form(None)):
    """Upload ricevuta spesa (pasto, altro)"""
    user = await get_current_user(request)
    
    rimborso = await db.rimborsi.find_one({"_id": ObjectId(rimborso_id)})
    if not rimborso:
        raise HTTPException(status_code=404, detail="Rimborso non trovato")
    
    if rimborso["user_id"] != user["id"] and not user_has_any_role(user, ["admin", "superadmin"]):
        raise HTTPException(status_code=403, detail="Non autorizzato")
    
    if file.content_type not in ["application/pdf", "image/jpeg", "image/png", "image/jpg"]:
        raise HTTPException(status_code=400, detail="Formato file non supportato")
    
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File troppo grande. Max 5MB")
    
    file_id = str(uuid.uuid4())
    ext = file.filename.split(".")[-1] if "." in file.filename else "pdf"
    filename = f"spesa_{file_id}.{ext}"
    filepath = UPLOAD_DIR / filename
    
    async with aiofiles.open(filepath, "wb") as f:
        await f.write(content)
    
    ricevuta_spesa = {
        "id": file_id,
        "filename": file.filename,
        "path": filename,
        "tipo": tipo,
        "descrizione": descrizione,
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.rimborsi.update_one(
        {"_id": ObjectId(rimborso_id)},
        {"$push": {"ricevute_spese": ricevuta_spesa}}
    )
    
    return ricevuta_spesa

@api_router.put("/rimborsi/{rimborso_id}")
async def update_rimborso(rimborso_id: str, rimborso_data: RimborsoUpdate, request: Request):
    user = await get_current_user(request)
    
    if not user_has_any_role(user, ["admin", "cassiere", "superadmin"]):
        raise HTTPException(status_code=403, detail="Solo admin/cassiere può gestire i rimborsi")
    
    rimborso = await db.rimborsi.find_one({"_id": ObjectId(rimborso_id)})
    if not rimborso:
        raise HTTPException(status_code=404, detail="Rimborso non trovato")
    
    if user_has_any_role(user, ["admin", "cassiere"]) and rimborso.get("sede_id") != user.get("sede_id"):
        raise HTTPException(status_code=403, detail="Non autorizzato per questa sede")
    
    # Lo stato "pagato" può essere impostato SOLO via /contabile (con upload obbligatorio)
    if rimborso_data.stato == "pagato":
        raise HTTPException(
            status_code=400,
            detail="Per pagare un rimborso devi caricare la contabile tramite l'apposito endpoint /contabile"
        )
    
    update_data = {k: v for k, v in rimborso_data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Nessun dato da aggiornare")
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.rimborsi.update_one({"_id": ObjectId(rimborso_id)}, {"$set": update_data})
    
    # Notifica all'utente richiedente per approvato/rifiutato
    stato_msg = {
        "approvato": "approvata",
        "rifiutato": "rifiutata",
    }
    if rimborso_data.stato and rimborso_data.stato in stato_msg:
        await _notify_user(
            user_id=rimborso["user_id"],
            notifica_data={
                "sede_id": rimborso.get("sede_id"),
                "tipo": "rimborso",
                "titolo": f"Richiesta rimborso {stato_msg[rimborso_data.stato]}",
                "messaggio": f"La tua richiesta di rimborso del {rimborso['data']} è stata {stato_msg[rimborso_data.stato]}",
                "rimborso_id": rimborso_id,
            }
        )
        await _log_audit(
            actor=user,
            action=f"rimborso.{rimborso_data.stato}",
            target_type="rimborso",
            target_id=rimborso_id,
            target_label=f"Rimborso del {rimborso['data']} - €{rimborso.get('importo_totale', 0):.2f}",
            sede_id=rimborso.get("sede_id"),
            old_value=rimborso.get("stato"),
            new_value=rimborso_data.stato,
        )
    
    return {"message": "Rimborso aggiornato"}

@api_router.post("/rimborsi/{rimborso_id}/contabile")
async def upload_contabile(rimborso_id: str, request: Request, file: UploadFile = File(...)):
    user = await get_current_user(request)
    
    if not user_has_any_role(user, ["admin", "cassiere", "superadmin"]):
        raise HTTPException(status_code=403, detail="Solo admin/cassiere può caricare contabili")
    
    rimborso = await db.rimborsi.find_one({"_id": ObjectId(rimborso_id)})
    if not rimborso:
        raise HTTPException(status_code=404, detail="Rimborso non trovato")
    
    if user_has_any_role(user, ["admin", "cassiere"]) and rimborso.get("sede_id") != user.get("sede_id"):
        raise HTTPException(status_code=403, detail="Non autorizzato per questa sede")
    
    if rimborso.get("stato") == "rifiutato":
        raise HTTPException(status_code=400, detail="Impossibile pagare un rimborso rifiutato")
    
    if rimborso.get("stato") == "pagato":
        raise HTTPException(status_code=400, detail="Rimborso già pagato")
    
    # Validazione file
    if file.content_type not in ["application/pdf", "image/jpeg", "image/png", "image/jpg"]:
        raise HTTPException(status_code=400, detail="Formato file non supportato (solo PDF, JPG, PNG)")
    
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File troppo grande. Max 5MB")
    
    file_id = str(uuid.uuid4())
    ext = file.filename.split(".")[-1] if "." in file.filename else "pdf"
    filename = f"contabile_{file_id}.{ext}"
    filepath = UPLOAD_DIR / filename
    
    async with aiofiles.open(filepath, "wb") as f:
        await f.write(content)
    
    # Se rimborso in_attesa, salta direttamente a pagato (auto-approvazione + pagamento)
    pagamento_diretto = rimborso.get("stato") == "in_attesa"
    
    await db.rimborsi.update_one(
        {"_id": ObjectId(rimborso_id)},
        {"$set": {
            "stato": "pagato",
            "contabile": {"filename": file.filename, "path": filename},
            "pagato_at": datetime.now(timezone.utc).isoformat(),
            "pagato_by": user["id"],
            "pagato_by_nome": f"{user['nome']} {user['cognome']}"
        }}
    )
    
    # Notifica all'utente richiedente
    if pagamento_diretto:
        msg_user = f"Il tuo rimborso del {rimborso['data']} è stato approvato e pagato. Contabile disponibile."
        titolo_user = "Rimborso approvato e pagato"
    else:
        msg_user = f"Il tuo rimborso del {rimborso['data']} è stato pagato. Contabile disponibile."
        titolo_user = "Rimborso pagato"
    
    await _notify_user(
        user_id=rimborso["user_id"],
        notifica_data={
            "sede_id": rimborso.get("sede_id"),
            "tipo": "rimborso",
            "titolo": titolo_user,
            "messaggio": msg_user,
            "rimborso_id": rimborso_id,
        }
    )
    
    # Notifica anche Admin + Cassiere della sede (escluso chi ha appena pagato)
    notifiche_admin = []
    async for u in db.users.find(
        {"ruolo": {"$in": ["admin", "cassiere"]}, "sede_id": rimborso.get("sede_id"), "_id": {"$ne": ObjectId(user["id"])}},
        {"_id": 1}
    ):
        notifiche_admin.append({
            "user_id": str(u["_id"]),
            "sede_id": rimborso.get("sede_id"),
            "tipo": "rimborso",
            "titolo": "Rimborso pagato",
            "messaggio": f"{user['nome']} {user['cognome']} ha pagato il rimborso del {rimborso['data']}",
            "rimborso_id": rimborso_id,
            "letto": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    if notifiche_admin:
        await db.notifiche.insert_many(notifiche_admin)
    
    await _log_audit(
        actor=user,
        action="rimborso.pay_direct" if pagamento_diretto else "rimborso.pay",
        target_type="rimborso",
        target_id=rimborso_id,
        target_label=f"Rimborso del {rimborso['data']} - €{rimborso.get('importo_totale', 0):.2f}",
        sede_id=rimborso.get("sede_id"),
        old_value=rimborso.get("stato"),
        new_value="pagato",
        note=f"Contabile caricata: {file.filename}",
    )
    
    return {"message": "Contabile caricata e rimborso pagato"}

# ==================== ANNUNCI (BULLETIN BOARD) ROUTES ====================

@api_router.get("/annunci")
async def get_annunci(request: Request):
    user = await get_current_user(request)
    
    query = {}
    if not user_has_any_role(user, ["superadmin", "superuser"]):
        query["$or"] = [
            {"sede_id": user.get("sede_id")},
            {"sede_id": None}
        ]
    
    annunci = []
    async for annuncio in db.annunci.find(query).sort("created_at", -1).limit(50):
        annuncio["id"] = str(annuncio["_id"])
        annuncio.pop("_id")
        annunci.append(annuncio)
    
    return annunci

@api_router.post("/annunci")
async def create_annuncio(
    request: Request,
    titolo: str = Form(...),
    contenuto: str = Form(...),
    link_documento: Optional[str] = Form(None),
    file: Optional[UploadFile] = File(None)
):
    user = await get_current_user(request)
    
    if not user_has_any_role(user, ["segreteria", "segretario", "admin", "superadmin"]):
        raise HTTPException(status_code=403, detail="Permessi insufficienti")
    
    allegato_filename = None
    allegato_path = None
    
    # Gestione upload file opzionale
    if file and file.filename:
        if file.content_type not in ["application/pdf", "image/jpeg", "image/png", "image/jpg"]:
            raise HTTPException(status_code=400, detail="Formato file non supportato (solo PDF, JPG, PNG)")
        
        content = await file.read()
        if len(content) > 5 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="File troppo grande. Max 5MB")
        
        file_id = str(uuid.uuid4())
        ext = file.filename.split(".")[-1] if "." in file.filename else "pdf"
        stored_name = f"annuncio_{file_id}.{ext}"
        filepath = UPLOAD_DIR / stored_name
        
        async with aiofiles.open(filepath, "wb") as f:
            await f.write(content)
        
        allegato_filename = file.filename
        allegato_path = stored_name
    
    annuncio_doc = {
        "titolo": titolo,
        "contenuto": contenuto,
        "link_documento": link_documento if link_documento else None,
        "allegato_filename": allegato_filename,
        "allegato_path": allegato_path,
        "sede_id": user.get("sede_id") if not user_has_role(user, "superadmin") else None,
        "autore_id": user["id"],
        "autore_nome": f"{user['nome']} {user['cognome']}",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    result = await db.annunci.insert_one(annuncio_doc)
    annuncio_doc["id"] = str(result.inserted_id)
    annuncio_doc.pop("_id", None)
    
    # Notifica TUTTI gli utenti della sede (o globali se sede_id None)
    await _notify_all_in_sede(
        sede_id=annuncio_doc["sede_id"],
        notifica_data={
            "sede_id": annuncio_doc["sede_id"],
            "tipo": "annuncio",
            "titolo": "Nuovo comunicato in bacheca",
            "messaggio": f"{annuncio_doc['autore_nome']}: {titolo}",
            "annuncio_id": annuncio_doc["id"],
        },
        exclude_user_id=user["id"]
    )
    
    return annuncio_doc

@api_router.get("/annunci/{annuncio_id}/download")
async def download_allegato_annuncio(annuncio_id: str, request: Request):
    user = await get_current_user(request)
    
    annuncio = await db.annunci.find_one({"_id": ObjectId(annuncio_id)})
    if not annuncio:
        raise HTTPException(status_code=404, detail="Annuncio non trovato")
    
    if not annuncio.get("allegato_path"):
        raise HTTPException(status_code=404, detail="Nessun allegato per questo annuncio")
    
    # Controllo sede per utenti non super
    if not user_has_any_role(user, ["superadmin", "superuser"]):
        if annuncio.get("sede_id") and annuncio["sede_id"] != user.get("sede_id"):
            raise HTTPException(status_code=403, detail="Non autorizzato")
    
    filepath = UPLOAD_DIR / annuncio["allegato_path"]
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="File non trovato")
    
    return FileResponse(filepath, filename=annuncio.get("allegato_filename", "allegato"))

@api_router.delete("/annunci/{annuncio_id}")
async def delete_annuncio(annuncio_id: str, request: Request):
    user = await get_current_user(request)
    
    if not user_has_any_role(user, ["segreteria", "segretario", "admin", "superadmin"]):
        raise HTTPException(status_code=403, detail="Permessi insufficienti")
    
    annuncio = await db.annunci.find_one({"_id": ObjectId(annuncio_id)})
    if not annuncio:
        raise HTTPException(status_code=404, detail="Annuncio non trovato")
    
    # Rimuovi il file fisico se presente
    if annuncio.get("allegato_path"):
        filepath = UPLOAD_DIR / annuncio["allegato_path"]
        if filepath.exists():
            filepath.unlink()
    
    await db.annunci.delete_one({"_id": ObjectId(annuncio_id)})
    
    return {"message": "Annuncio eliminato"}

# ==================== CONTATTI / LINK SIDEBAR ROUTES ====================

VALID_CONTATTO_TIPI = ["link", "whatsapp", "telegram", "email", "telefono"]
EDIT_CONTATTO_ROLES = ["admin", "segretario", "segreteria", "superadmin"]


@api_router.get("/contatti")
async def get_contatti(request: Request):
    """Restituisce i contatti della sede dell'utente. SuperAdmin/SuperUser vede tutti."""
    user = await get_current_user(request)
    
    query = {}
    if not user_has_any_role(user, ["superadmin", "superuser"]):
        query["sede_id"] = user.get("sede_id")
    
    contatti = []
    async for c in db.contatti.find(query).sort("ordine", 1):
        c["id"] = str(c["_id"])
        c.pop("_id")
        contatti.append(c)
    
    return contatti


@api_router.post("/contatti")
async def create_contatto(contatto_data: ContattoCreate, request: Request):
    user = await get_current_user(request)
    
    if not user_has_any_role(user, EDIT_CONTATTO_ROLES):
        raise HTTPException(status_code=403, detail="Permessi insufficienti")
    
    if contatto_data.tipo not in VALID_CONTATTO_TIPI:
        raise HTTPException(status_code=400, detail=f"Tipo non valido. Usa: {', '.join(VALID_CONTATTO_TIPI)}")
    
    contatto_doc = {
        "titolo": contatto_data.titolo,
        "descrizione": contatto_data.descrizione,
        "tipo": contatto_data.tipo,
        "valore": contatto_data.valore,
        "sede_id": user.get("sede_id"),
        "ordine": 0,
        "created_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    result = await db.contatti.insert_one(contatto_doc)
    contatto_doc["id"] = str(result.inserted_id)
    contatto_doc.pop("_id", None)
    
    return contatto_doc


@api_router.put("/contatti/{contatto_id}")
async def update_contatto(contatto_id: str, contatto_data: ContattoUpdate, request: Request):
    user = await get_current_user(request)
    
    if not user_has_any_role(user, EDIT_CONTATTO_ROLES):
        raise HTTPException(status_code=403, detail="Permessi insufficienti")
    
    contatto = await db.contatti.find_one({"_id": ObjectId(contatto_id)})
    if not contatto:
        raise HTTPException(status_code=404, detail="Contatto non trovato")
    
    # Solo stessa sede (eccetto superadmin)
    if not user_has_role(user, "superadmin") and contatto.get("sede_id") != user.get("sede_id"):
        raise HTTPException(status_code=403, detail="Non autorizzato per questa sede")
    
    update_data = {k: v for k, v in contatto_data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Nessun dato da aggiornare")
    
    if "tipo" in update_data and update_data["tipo"] not in VALID_CONTATTO_TIPI:
        raise HTTPException(status_code=400, detail=f"Tipo non valido. Usa: {', '.join(VALID_CONTATTO_TIPI)}")
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.contatti.update_one({"_id": ObjectId(contatto_id)}, {"$set": update_data})
    
    return {"message": "Contatto aggiornato"}


@api_router.delete("/contatti/{contatto_id}")
async def delete_contatto(contatto_id: str, request: Request):
    user = await get_current_user(request)
    
    if not user_has_any_role(user, EDIT_CONTATTO_ROLES):
        raise HTTPException(status_code=403, detail="Permessi insufficienti")
    
    contatto = await db.contatti.find_one({"_id": ObjectId(contatto_id)})
    if not contatto:
        raise HTTPException(status_code=404, detail="Contatto non trovato")
    
    if not user_has_role(user, "superadmin") and contatto.get("sede_id") != user.get("sede_id"):
        raise HTTPException(status_code=403, detail="Non autorizzato per questa sede")
    
    await db.contatti.delete_one({"_id": ObjectId(contatto_id)})
    
    return {"message": "Contatto eliminato"}



# ==================== DOCUMENTI ROUTES ====================

@api_router.get("/documenti")
async def get_documenti(request: Request, categoria: Optional[str] = None):
    user = await get_current_user(request)
    
    query = {}
    if not user_has_any_role(user, ["superadmin", "superuser"]):
        query["$or"] = [
            {"sede_id": user.get("sede_id")},
            {"sede_id": None}
        ]
    
    if categoria:
        query["categoria"] = categoria
    
    documenti = []
    async for doc in db.documenti.find(query).sort("created_at", -1):
        doc["id"] = str(doc["_id"])
        doc.pop("_id")
        documenti.append(doc)
    
    return documenti

@api_router.post("/documenti")
async def upload_documento(
    request: Request,
    file: UploadFile = File(...),
    nome: str = Form(...),
    categoria: str = Form(...),
    descrizione: str = Form(None)
):
    user = await get_current_user(request)
    
    if not user_has_any_role(user, ["segreteria", "segretario", "admin", "superadmin"]):
        raise HTTPException(status_code=403, detail="Permessi insufficienti")
    
    if file.content_type not in ["application/pdf", "image/jpeg", "image/png", "image/jpg"]:
        raise HTTPException(status_code=400, detail="Formato file non supportato")
    
    content = await file.read()
    if len(content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File troppo grande. Max 5MB")
    
    file_id = str(uuid.uuid4())
    ext = file.filename.split(".")[-1] if "." in file.filename else "pdf"
    filename = f"doc_{file_id}.{ext}"
    filepath = UPLOAD_DIR / filename
    
    async with aiofiles.open(filepath, "wb") as f:
        await f.write(content)
    
    doc_record = {
        "nome": nome,
        "categoria": categoria,
        "descrizione": descrizione,
        "filename": file.filename,
        "path": filename,
        "sede_id": user.get("sede_id") if not user_has_role(user, "superadmin") else None,
        "uploaded_by": user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    result = await db.documenti.insert_one(doc_record)
    doc_record["id"] = str(result.inserted_id)
    doc_record.pop("_id", None)
    
    # Notifica TUTTI gli utenti della sede (o globali se sede_id None)
    await _notify_all_in_sede(
        sede_id=doc_record["sede_id"],
        notifica_data={
            "sede_id": doc_record["sede_id"],
            "tipo": "documento",
            "titolo": "Nuovo documento disponibile",
            "messaggio": f"{user['nome']} {user['cognome']} ha caricato: {nome}",
            "documento_id": doc_record["id"],
        },
        exclude_user_id=user["id"]
    )
    
    return doc_record

@api_router.get("/documenti/{doc_id}/download")
async def download_documento(doc_id: str, request: Request):
    user = await get_current_user(request)
    
    doc = await db.documenti.find_one({"_id": ObjectId(doc_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Documento non trovato")
    
    if not user_has_any_role(user, ["superadmin", "superuser"]):
        if doc.get("sede_id") and doc["sede_id"] != user.get("sede_id"):
            raise HTTPException(status_code=403, detail="Non autorizzato")
    
    filepath = UPLOAD_DIR / doc["path"]
    if not filepath.exists():
        raise HTTPException(status_code=404, detail="File non trovato")
    
    return FileResponse(filepath, filename=doc["filename"])

@api_router.delete("/documenti/{doc_id}")
async def delete_documento(doc_id: str, request: Request):
    user = await get_current_user(request)
    
    if not user_has_any_role(user, ["segreteria", "segretario", "admin", "superadmin"]):
        raise HTTPException(status_code=403, detail="Permessi insufficienti")
    
    doc = await db.documenti.find_one({"_id": ObjectId(doc_id)})
    if not doc:
        raise HTTPException(status_code=404, detail="Documento non trovato")
    
    filepath = UPLOAD_DIR / doc["path"]
    if filepath.exists():
        filepath.unlink()
    
    await db.documenti.delete_one({"_id": ObjectId(doc_id)})
    
    return {"message": "Documento eliminato"}

# ==================== NOTIFICHE ROUTES ====================

@api_router.get("/notifiche")
async def get_notifiche(request: Request):
    user = await get_current_user(request)
    
    # Notifiche dirette all'utente + retrocompatibilità con notifiche vecchie (user_id=None, sede_id specifica)
    query = {"$or": [
        {"user_id": user["id"]},
        {"user_id": None, "sede_id": user.get("sede_id")}
    ]}
    
    notifiche = []
    async for notifica in db.notifiche.find(query).sort("created_at", -1).limit(50):
        notifica["id"] = str(notifica["_id"])
        notifica.pop("_id")
        notifiche.append(notifica)
    
    return notifiche

@api_router.put("/notifiche/{notifica_id}/letto")
async def mark_notifica_letta(notifica_id: str, request: Request):
    await get_current_user(request)  # solo per autenticare
    
    await db.notifiche.update_one(
        {"_id": ObjectId(notifica_id)},
        {"$set": {"letto": True}}
    )
    
    return {"message": "Notifica segnata come letta"}

@api_router.put("/notifiche/letto-tutte")
async def mark_all_notifiche_lette(request: Request):
    user = await get_current_user(request)
    
    query = {"$or": [
        {"user_id": user["id"]},
        {"user_id": None, "sede_id": user.get("sede_id")}
    ]}
    
    await db.notifiche.update_many(query, {"$set": {"letto": True}})
    
    return {"message": "Tutte le notifiche segnate come lette"}

# ==================== USERS MANAGEMENT ====================

@api_router.get("/users")
async def get_users(request: Request):
    user = await get_current_user(request)
    
    if not user_has_any_role(user, ["admin", "cassiere", "superadmin", "superuser", "segretario"]):
        raise HTTPException(status_code=403, detail="Permessi insufficienti")
    
    query = {}
    if not user_has_any_role(user, ["superadmin", "superuser"]):
        query["sede_id"] = user.get("sede_id")
    
    users = []
    async for u in db.users.find(query, {"password_hash": 0}):
        u["id"] = str(u["_id"])
        u.pop("_id")
        # Multi-ruolo: garantisci campo `ruoli` sempre presente
        if not u.get("ruoli"):
            u["ruoli"] = [u["ruolo"]] if u.get("ruolo") else []
        if u.get("sede_id"):
            sede = await db.sedi.find_one({"_id": ObjectId(u["sede_id"])})
            if sede:
                u["sede_nome"] = sede["nome"]
        users.append(u)
    
    return users

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, user_data: UserUpdate, request: Request):
    current_user = await get_current_user(request)
    
    if user_id != current_user["id"] and not user_has_any_role(current_user, ["admin", "superadmin"]):
        raise HTTPException(status_code=403, detail="Permessi insufficienti")
    
    update_data = {k: v for k, v in user_data.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Nessun dato da aggiornare")
    
    await db.users.update_one({"_id": ObjectId(user_id)}, {"$set": update_data})
    
    return {"message": "Utente aggiornato"}


@api_router.put("/users/{user_id}/ruolo")
async def update_user_role(user_id: str, request: Request, payload: UpdateRuoliRequest):
    """
    Aggiorna i ruoli di un utente.
    Accetta una lista `ruoli` (multi-ruolo).
    Vincoli:
    - 'iscritto' non combinabile con altri ruoli
    - Solo superadmin può assegnare/togliere superuser/superadmin
    """
    current_user = await get_current_user(request)
    
    if not user_has_any_role(current_user, ["admin", "superadmin"]):
        raise HTTPException(status_code=403, detail="Permessi insufficienti")
    
    # Determina i ruoli assegnabili in base al ruolo del current_user
    valid_roles = ["iscritto", "delegato", "segreteria", "segretario", "cassiere", "admin"]
    if user_has_role(current_user, "superadmin"):
        valid_roles.extend(["superuser", "superadmin"])
    
    # Valida ogni ruolo richiesto sia tra quelli assegnabili
    for r in payload.ruoli:
        if r not in valid_roles:
            raise HTTPException(status_code=400, detail=f"Ruolo '{r}' non assegnabile")
    
    # Normalizza (dedup + iscritto exclusive)
    nuovi_ruoli = normalize_roles_input(payload.ruoli, None)
    
    target_user = await db.users.find_one({"_id": ObjectId(user_id)})
    if not target_user:
        raise HTTPException(status_code=404, detail="Utente non trovato")
    
    # Aggiorna sia il campo legacy `ruolo` (= primario) sia `ruoli` (array completo)
    await db.users.update_one(
        {"_id": ObjectId(user_id)},
        {"$set": {"ruolo": nuovi_ruoli[0], "ruoli": nuovi_ruoli}}
    )
    
    old_ruoli = target_user.get("ruoli") or [target_user.get("ruolo")]
    if sorted(old_ruoli) != sorted(nuovi_ruoli):
        await _log_audit(
            actor=current_user,
            action="user.change_role",
            target_type="user",
            target_id=user_id,
            target_label=f"{target_user.get('nome', '')} {target_user.get('cognome', '')} ({target_user.get('email', '')})".strip(),
            sede_id=target_user.get("sede_id"),
            old_value=", ".join(old_ruoli),
            new_value=", ".join(nuovi_ruoli),
        )
    
    return {"message": "Ruoli aggiornati", "ruoli": nuovi_ruoli}

@api_router.post("/system/check-pending-reimbursements")
async def trigger_pending_check(request: Request):
    """Lancia manualmente il check rimborsi pendenti >7gg (solo superadmin/admin).
    Utile per testing e debug. In produzione lo scheduler automatico gira ogni notte."""
    user = await get_current_user(request)
    if not user_has_any_role(user, ["superadmin", "admin"]):
        raise HTTPException(status_code=403, detail="Permessi insufficienti")
    result = await _check_pending_reimbursements()
    return result


# ==================== REPORTS ====================

@api_router.get("/reports/rimborsi-annuali")
async def get_report_rimborsi_annuali(request: Request, anno: int):
    user = await get_current_user(request)
    
    if not user_has_any_role(user, ["admin", "cassiere", "superadmin", "superuser"]):
        raise HTTPException(status_code=403, detail="Permessi insufficienti")
    
    query = {"data": {"$regex": f"^{anno}"}}
    if not user_has_any_role(user, ["superadmin", "superuser"]):
        query["sede_id"] = user.get("sede_id")
    
    pipeline = [
        {"$match": query},
        {"$group": {
            "_id": "$user_id",
            "totale_rimborsi": {"$sum": 1},
            "totale_importo": {"$sum": "$importo_totale"},
            "totale_km": {"$sum": "$km_totali"},
            "rimborsi_pagati": {"$sum": {"$cond": [{"$eq": ["$stato", "pagato"]}, 1, 0]}},
            "importo_pagato": {"$sum": {"$cond": [{"$eq": ["$stato", "pagato"]}, "$importo_totale", 0]}}
        }}
    ]
    
    results = []
    async for result in db.rimborsi.aggregate(pipeline):
        user_doc = await db.users.find_one({"_id": ObjectId(result["_id"])})
        if user_doc:
            result["user_nome"] = f"{user_doc['nome']} {user_doc['cognome']}"
            result["user_email"] = user_doc["email"]
            result["user_iban"] = user_doc.get("iban", "")
        results.append(result)
    
    return results

@api_router.get("/reports/rimborsi-export")
async def export_rimborsi(request: Request, anno: int, formato: str = "csv"):
    user = await get_current_user(request)
    
    if not user_has_any_role(user, ["admin", "cassiere", "superadmin", "superuser"]):
        raise HTTPException(status_code=403, detail="Permessi insufficienti")
    
    query = {"data": {"$regex": f"^{anno}"}}
    if not user_has_any_role(user, ["superadmin", "superuser"]):
        query["sede_id"] = user.get("sede_id")
    
    rimborsi = []
    async for rimborso in db.rimborsi.find(query).sort("data", 1):
        rimborso_user = await db.users.find_one({"_id": ObjectId(rimborso["user_id"])})
        motivo = await db.motivi_rimborso.find_one({"_id": ObjectId(rimborso["motivo_id"])}) if rimborso.get("motivo_id") else None
        
        rimborsi.append({
            "Data": rimborso["data"],
            "Utente": f"{rimborso_user['nome']} {rimborso_user['cognome']}" if rimborso_user else "N/A",
            "Email": rimborso_user["email"] if rimborso_user else "",
            "IBAN": rimborso_user.get("iban", "") if rimborso_user else "",
            "Motivo": motivo["nome"] if motivo else "N/A",
            "Partenza": rimborso["indirizzo_partenza"],
            "Arrivo": rimborso["indirizzo_arrivo"],
            "KM Totali": rimborso["km_totali"],
            "Importo KM": f"{rimborso['importo_km']:.2f}",
            "Importo Pasti": f"{rimborso.get('importo_pasti', 0):.2f}",
            "Autostrada": f"{rimborso.get('costo_autostrada', 0):.2f}",
            "Totale": f"{rimborso['importo_totale']:.2f}",
            "Stato": rimborso["stato"],
            "Note": rimborso.get("note", "")
        })
    
    if formato == "csv":
        output = io.StringIO()
        if rimborsi:
            writer = csv.DictWriter(output, fieldnames=rimborsi[0].keys())
            writer.writeheader()
            writer.writerows(rimborsi)
        
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=rimborsi_{anno}.csv"}
        )
    elif formato == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Rimborsi {anno}"
        
        headers = list(rimborsi[0].keys()) if rimborsi else [
            "Data", "Utente", "Email", "IBAN", "Motivo", "Partenza", "Arrivo",
            "KM Totali", "Importo KM", "Importo Pasti", "Autostrada", "Totale", "Stato", "Note"
        ]
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E4D8C", end_color="1E4D8C", fill_type="solid")
        thin = Side(border_style="thin", color="CCCCCC")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        for row_idx, r in enumerate(rimborsi, start=2):
            for col_idx, key in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=r.get(key, ""))
                cell.border = border
                cell.alignment = Alignment(vertical="center")
        
        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
        
        # Header altezza
        ws.row_dimensions[1].height = 25
        
        # Riga totale
        if rimborsi:
            total = sum(float(r["Totale"]) for r in rimborsi)
            total_row = len(rimborsi) + 2
            ws.cell(row=total_row, column=1, value="TOTALE").font = Font(bold=True)
            cell = ws.cell(row=total_row, column=headers.index("Totale") + 1, value=f"{total:.2f}")
            cell.font = Font(bold=True, color="1E4D8C")
        
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return StreamingResponse(
            iter([out.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=rimborsi_{anno}.xlsx"}
        )
    elif formato == "pdf":
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import mm
        
        out = io.BytesIO()
        doc = SimpleDocTemplate(
            out, pagesize=landscape(A4),
            rightMargin=10*mm, leftMargin=10*mm, topMargin=15*mm, bottomMargin=15*mm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        title_style = styles["Title"]
        title_style.textColor = colors.HexColor("#1E4D8C")
        
        # Titolo
        sede_nome = ""
        if user.get("sede_id"):
            sede = await db.sedi.find_one({"_id": ObjectId(user["sede_id"])})
            if sede:
                sede_nome = f" - {sede['nome']}"
        
        elements.append(Paragraph(f"Rendiconto Rimborsi {anno}{sede_nome}", title_style))
        elements.append(Paragraph(
            f"<font size=9 color='#666'>Generato da {user.get('nome', '')} {user.get('cognome', '')} il {datetime.now().strftime('%d/%m/%Y %H:%M')}</font>",
            styles["Normal"]
        ))
        elements.append(Spacer(1, 8*mm))
        
        # Tabella (riduco colonne per A4 landscape)
        pdf_headers = ["Data", "Utente", "Motivo", "KM", "Importo €", "Stato"]
        data = [pdf_headers]
        total = 0.0
        for r in rimborsi:
            data.append([
                r["Data"],
                r["Utente"],
                r["Motivo"],
                str(r["KM Totali"]),
                r["Totale"],
                r["Stato"].upper()
            ])
            try:
                total += float(r["Totale"])
            except (ValueError, TypeError):
                pass
        
        # Riga totale
        data.append(["", "", "", "TOTALE", f"{total:.2f}", ""])
        
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E4D8C")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor("#F5F7FA")]),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#FFF7E0")),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
        
        doc.build(elements)
        out.seek(0)
        return StreamingResponse(
            iter([out.getvalue()]),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=rimborsi_{anno}.pdf"}
        )
    else:
        return rimborsi

# ==================== STARTUP ====================

@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.sedi.create_index("codice", unique=True)
    await db.login_attempts.create_index("identifier")
    
    # === MIGRAZIONE MULTI-RUOLO ===
    # Per ogni utente senza il campo `ruoli`, lo crea a partire da `ruolo` legacy.
    migrated = 0
    async for u in db.users.find({"ruoli": {"$exists": False}}, {"_id": 1, "ruolo": 1}):
        if u.get("ruolo"):
            await db.users.update_one(
                {"_id": u["_id"]},
                {"$set": {"ruoli": [u["ruolo"]]}}
            )
            migrated += 1
    if migrated:
        logger.info(f"Migrazione multi-ruolo: {migrated} utenti aggiornati con campo 'ruoli'.")
    
    admin_email = os.environ.get("ADMIN_EMAIL", "superadmin@sla.it")
    admin_password = os.environ.get("ADMIN_PASSWORD", "SlaAdmin2024!")
    
    existing = await db.users.find_one({"email": admin_email})
    if existing is None:
        hashed = hash_password(admin_password)
        await db.users.insert_one({
            "email": admin_email,
            "password_hash": hashed,
            "nome": "Super",
            "cognome": "Admin",
            "ruolo": "superadmin",
            "ruoli": ["superadmin"],
            "sede_id": None,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info(f"SuperAdmin creato: {admin_email}")
    elif not verify_password(admin_password, existing["password_hash"]):
        await db.users.update_one(
            {"email": admin_email},
            {"$set": {"password_hash": hash_password(admin_password)}}
        )
        logger.info("Password SuperAdmin aggiornata")
    
    # Seed default motivi rimborso with richiede_note flag
    motivi_default = [
        {"nome": "RSA", "richiede_note": False},
        {"nome": "Sede", "richiede_note": False},
        {"nome": "Altro", "richiede_note": True}
    ]
    for motivo in motivi_default:
        existing_motivo = await db.motivi_rimborso.find_one({"nome": motivo["nome"]})
        if not existing_motivo:
            await db.motivi_rimborso.insert_one({
                "nome": motivo["nome"],
                "descrizione": None,
                "richiede_note": motivo["richiede_note"],
                "created_at": datetime.now(timezone.utc).isoformat()
            })
    
    # Remove old motivi that are not in the new list
    await db.motivi_rimborso.delete_many({"nome": {"$nin": ["RSA", "Sede", "Altro"]}})
    
    # Only keep A22 sede
    await db.sedi.delete_many({"codice": {"$nin": ["A22"]}})
    
    # Ensure A22 exists
    existing_a22 = await db.sedi.find_one({"codice": "A22"})
    if not existing_a22:
        await db.sedi.insert_one({
            "nome": "Autostrada del Brennero",
            "codice": "A22",
            "indirizzo": None,
            "tariffa_km": 0.35,
            "rimborso_pasti": 15.0,
            "rimborso_autostrada": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
    
    creds_path = Path("/app/memory/test_credentials.md")
    creds_path.parent.mkdir(exist_ok=True)
    creds_path.write_text(f"""# Test Credentials

## SuperAdmin
- Email: {admin_email}
- Password: {admin_password}
- Ruolo: superadmin

## Auth Endpoints
- POST /api/auth/register
- POST /api/auth/login
- POST /api/auth/logout
- GET /api/auth/me
- POST /api/auth/refresh

## Sedi
- A22 - Autostrada del Brennero

## Motivi Rimborso
- RSA (note non obbligatorie)
- Sede (note non obbligatorie)
- Altro (note OBBLIGATORIE)
""")
    
    logger.info("Database inizializzato")
    
    # === SCHEDULER PROMEMORIA RIMBORSI PENDENTI >7gg ===
    # Lancia il task in background che ogni notte alle 08:00 controlla
    # i rimborsi in_attesa / approvato fermi da >=7 giorni e notifica i cassieri.
    asyncio.create_task(_pending_reimbursements_scheduler())
    logger.info("Scheduler promemoria rimborsi avviato")

@app.on_event("shutdown")
async def shutdown():
    client.close()

app.include_router(api_router)

# ==================== SECURITY HEADERS MIDDLEWARE ====================
# Aggiunge header HTTP standard per protezione contro XSS, clickjacking,
# MIME sniffing, downgrade HTTPS. Si applicano a TUTTE le risposte.

@app.middleware("http")
async def security_headers_middleware(request: Request, call_next):
    response = await call_next(request)
    # HSTS: forza HTTPS per 1 anno (browser ricordano e bloccano HTTP)
    response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    # Impedisce embed in iframe → anti-clickjacking
    response.headers["X-Frame-Options"] = "DENY"
    # Impedisce MIME-sniffing → riduce rischio XSS via upload
    response.headers["X-Content-Type-Options"] = "nosniff"
    # Limita info inviate ad altri siti via referer
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    # Disabilita feature browser non usate (camera, microfono, geo, ecc.)
    response.headers["Permissions-Policy"] = (
        "camera=(), microphone=(), geolocation=(), payment=(), usb=()"
    )
    # CSP: limita le origini di script/style/img consentite
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' 'unsafe-eval' https://maps.googleapis.com; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com data:; "
        "img-src 'self' data: blob: https://*.googleapis.com https://*.gstatic.com; "
        "connect-src 'self' https://maps.googleapis.com; "
        "frame-ancestors 'none'; "
        "object-src 'none'; "
        "base-uri 'self'"
    )
    return response


app.add_middleware(
    CORSMiddleware,
    allow_origin_regex=FRONTEND_ORIGIN_REGEX,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
